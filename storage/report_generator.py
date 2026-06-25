"""
Excel, PDF, and JSON report generation for live RFC-based analysis results.
"""
from __future__ import annotations
import io
import json
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape as rl_landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from core.domain_models import LiveApplicabilityResult, LiveSystemInfo, NoteMetadata

logger = logging.getLogger(__name__)

_HDR_FILL = "1F3864"
_STATUS_FILL = {
    "Applicable":           "FFCDD2",
    "Not Applicable":       "C8E6C9",
    "Already Implemented":  "B3E5FC",
    "Needs Manual Review":  "FFF9C4",
    "Insufficient Data":    "F3E5F5",
}


def _hdr(ws, values, row=1):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def export_excel(
    results: List[LiveApplicabilityResult],
    system: LiveSystemInfo,
    notes: List[NoteMetadata],
) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _hdr(ws_sum, ["KPI", "Value"])
    kpis = [
        ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("System SID", system.sid),
        ("Client", system.client),
        ("SAP Release", system.sap_release),
        ("Kernel", f"{system.kernel_release} PL{system.kernel_patch}"),
        ("Components Installed", len(system.components)),
        ("Notes Checked", len(results)),
        ("Applicable",          sum(1 for r in results if r.status == "Applicable")),
        ("Not Applicable",      sum(1 for r in results if r.status == "Not Applicable")),
        ("Already Implemented", sum(1 for r in results if r.status == "Already Implemented")),
        ("Needs Manual Review", sum(1 for r in results if r.status == "Needs Manual Review")),
        ("Insufficient Data",   sum(1 for r in results if r.status == "Insufficient Data")),
    ]
    for k, v in kpis:
        ws_sum.append([k, v])

    # Sheet 2: Results & Evidence
    ws_res = wb.create_sheet("Results & Evidence")
    headers = [
        "Note #", "Title", "Severity", "CVSS", "Status", "Confidence",
        "Component Checked", "Component Found", "Installed Release", "Required Release",
        "Installed SP", "Req SP From", "Req SP To", "SP In Range",
        "In CWBNTCUST", "PRSTATUS", "Kernel", "Decision Reason", "Action",
    ]
    _hdr(ws_res, headers)
    for r in results:
        ev = r.evidence
        row_idx = ws_res.max_row + 1
        ws_res.append([
            r.note_number, r.note_title, r.note_severity, r.note_cvss,
            r.status, f"{r.confidence:.0%}",
            ev.component.required_component,
            "Yes" if ev.component.component_found else "No",
            ev.component.installed_release, ev.component.required_release,
            ev.sp.installed_sp, ev.sp.required_sp_from, ev.sp.required_sp_to,
            ("Yes" if ev.sp.in_range else "No") if ev.sp.in_range is not None else "N/A",
            "Yes" if ev.implementation.note_in_cwbntcust else "No",
            ev.implementation.prstatus or "—",
            ev.kernel_release,
            ev.reason,
            r.recommended_action,
        ])
        fill = _STATUS_FILL.get(r.status, "FFFFFF")
        ws_res.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor=fill)

    # Sheet 3: System Inventory
    ws_inv = wb.create_sheet("System Inventory")
    _hdr(ws_inv, ["Component", "Release", "SP Level", "Patch Level", "Description"])
    for c in system.components:
        ws_inv.append([c.name, c.release, c.sp_level, c.patch_level, c.description])

    # Sheet 4: Note Metadata
    ws_notes = wb.create_sheet("Note Metadata")
    _hdr(ws_notes, ["Note #", "Title", "Severity", "CVSS", "Published", "Source",
                    "Components", "Matrix Rows", "Warnings"])
    for n in notes:
        ws_notes.append([
            n.note_number, n.title, n.severity, n.cvss_score,
            n.published_date, n.source,
            ", ".join(n.components), len(n.applicability_matrix),
            "; ".join(n.parser_warnings),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(results: List[LiveApplicabilityResult], system: LiveSystemInfo) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=rl_landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("SAP Security Note Applicability Report", styles["Title"]),
        Paragraph(
            f"System: <b>{system.sid}</b> | Client: {system.client} | "
            f"SAP Release: {system.sap_release} | "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles["Normal"],
        ),
        Spacer(1, 0.4*cm),
    ]

    data = [["Note #", "Title", "Severity", "Status", "SP Installed", "SP Range", "Decision Reason"]]
    for r in results:
        ev = r.evidence
        sp_range = f"{ev.sp.required_sp_from}–{ev.sp.required_sp_to}" if ev.sp.required_sp_from else "—"
        data.append([
            r.note_number,
            Paragraph(r.note_title[:55] or "—", styles["Normal"]),
            r.note_severity,
            r.status,
            ev.sp.installed_sp or "—",
            sp_range,
            Paragraph(ev.reason[:120] if ev.reason else "—", styles["Normal"]),
        ])

    tbl = LongTable(data, colWidths=[2*cm, 6*cm, 2*cm, 3.5*cm, 2*cm, 2*cm, 9*cm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(tbl)
    doc.build(story)
    return buf.getvalue()


def export_json(
    results: List[LiveApplicabilityResult],
    system: LiveSystemInfo,
    notes: List[NoteMetadata],
) -> bytes:
    payload = {
        "report_generated": datetime.now().isoformat(timespec="seconds"),
        "system": {
            "sid": system.sid, "client": system.client, "host": system.host,
            "sap_release": system.sap_release,
            "kernel": f"{system.kernel_release} PL{system.kernel_patch}",
        },
        "results": [_r2d(r) for r in results],
    }
    return json.dumps(payload, indent=2).encode("utf-8")


def _r2d(r: LiveApplicabilityResult) -> dict:
    ev = r.evidence
    return {
        "note_number": r.note_number, "title": r.note_title,
        "severity": r.note_severity, "cvss": r.note_cvss,
        "status": r.status, "confidence": r.confidence,
        "checked_at": r.checked_at, "action": r.recommended_action,
        "evidence": {
            "component": {
                "required": ev.component.required_component,
                "found": ev.component.component_found,
                "installed_release": ev.component.installed_release,
                "required_release": ev.component.required_release,
                "release_match": ev.component.release_match,
            },
            "sp": {
                "installed": ev.sp.installed_sp,
                "from": ev.sp.required_sp_from,
                "to": ev.sp.required_sp_to,
                "in_range": ev.sp.in_range,
            },
            "implementation": {
                "in_cwbntcust": ev.implementation.note_in_cwbntcust,
                "prstatus": ev.implementation.prstatus,
                "already_implemented": ev.implementation.already_implemented,
            },
            "reason": ev.reason,
        },
    }
