"""Enhanced Excel report with KPI dashboard, risk summary, and patch planner."""

from __future__ import annotations
import io
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

from core.domain_models import (
    ApplicabilityResult,
    Landscape,
    RiskSummary,
    SapSecurityNote,
)
from core.risk_analytics import build_patch_planner

_HEADER_FILL = "4472C4"
_STATUS_FILL = {
    "Applicable": "FFC7CE",
    "Not Applicable": "C6EFCE",
    "Needs Manual Review": "FFEB9C",
}


def _hdr(ws, values: list, row: int = 1) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def export_excel_report_future(
    landscape: Landscape,
    notes: List[SapSecurityNote],
    results: List[ApplicabilityResult],
    risk_summaries: List[RiskSummary],
) -> bytes:
    wb = openpyxl.Workbook()

    # ----- Sheet 1: KPI Dashboard -----
    ws_kpi = wb.active
    ws_kpi.title = "KPI Dashboard"
    total = len(results)
    applicable = sum(1 for r in results if r.status == "Applicable")
    not_app = sum(1 for r in results if r.status == "Not Applicable")
    manual = sum(1 for r in results if r.status == "Needs Manual Review")

    _hdr(ws_kpi, ["KPI", "Value"])
    kpis = [
        ("Total SIDs", len(landscape.systems)),
        ("Total Notes Imported", len(notes)),
        ("Total Evaluations", total),
        ("Applicable", applicable),
        ("Not Applicable", not_app),
        ("Needs Manual Review", manual),
        ("Critical Notes", sum(1 for n in notes if n.severity.lower() == "critical")),
        ("High Notes", sum(1 for n in notes if n.severity.lower() == "high")),
        ("Medium Notes", sum(1 for n in notes if n.severity.lower() == "medium")),
        ("Low Notes", sum(1 for n in notes if n.severity.lower() == "low")),
    ]
    for k, v in kpis:
        ws_kpi.append([k, v])

    # ----- Sheet 2: Landscape -----
    ws_land = wb.create_sheet("Landscape")
    _hdr(ws_land, ["SID", "Type", "Environment", "BASIS", "ABA", "Kernel", "Kernel Patch", "# Components"])
    for sys in landscape.systems:
        ws_land.append([
            sys.sid, sys.system_type, sys.environment,
            sys.sap_basis_release, sys.sap_aba_release,
            sys.kernel_release, sys.kernel_patch_level, len(sys.components),
        ])

    # ----- Sheet 3: Risk Summary -----
    ws_risk = wb.create_sheet("Risk Summary")
    _hdr(ws_risk, ["SID", "Critical", "High", "Medium", "Low", "Risk Score (0-100)",
                   "Exposure Score", "Avg Patch Age (days)"])
    for rs in risk_summaries:
        ws_risk.append([
            rs.sid, rs.critical_count, rs.high_count, rs.medium_count, rs.low_count,
            rs.risk_score, rs.exposure_score, rs.avg_patch_age_days,
        ])

    # ----- Sheet 4: Applicability Matrix -----
    ws_app = wb.create_sheet("Applicability Matrix")
    _hdr(ws_app, ["SID", "Note #", "Title", "Severity", "CVSS", "Status", "Confidence",
                  "Action", "Evidence", "Reasons"])
    note_map = {n.note_number: n for n in notes}
    for r in results:
        note = note_map.get(r.note_number)
        fill = _STATUS_FILL.get(r.status, "FFFFFF")
        row_idx = ws_app.max_row + 1
        ws_app.append([
            r.sid,
            r.note_number,
            note.title if note else "",
            note.severity if note else "",
            note.cvss_score if note else 0,
            r.status,
            f"{r.confidence:.0%}",
            r.recommended_action[:200] if r.recommended_action else "",
            r.evidence[:300] if r.evidence else "",
            "; ".join(f"{rr.flag}: {rr.reason}" for rr in r.reasons[:5]),
        ])
        ws_app.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor=fill)

    # ----- Sheet 5: Patch Plan -----
    ws_plan = wb.create_sheet("Patch Plan")
    _hdr(ws_plan, ["Priority", "SID", "Note #", "Title", "Severity", "Effort (1-5)", "Effort Rationale"])
    plan = build_patch_planner(notes, results)
    for i, row in enumerate(plan, 1):
        ws_plan.append([i, row["sid"], row["note_number"], row["title"],
                        row["severity"], row["effort_score"], row["effort_rationale"]])

    # ----- Sheet 6: Parser Warnings -----
    ws_warn = wb.create_sheet("Parser Warnings")
    _hdr(ws_warn, ["Note #", "Filename", "Warning"])
    for note in notes:
        for w in note.parser_warnings:
            ws_warn.append([note.note_number, note.title, w])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
